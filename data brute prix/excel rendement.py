import openpyxl as xl
import numpy as np


Prix_liste=[]

for j in range(12,19):
    data_j=xl.load_workbook(f'GUI_ENERGY_PRICES_202410{j-1}2200-202410{j}2200.xlsx',read_only=True)
    data_j_s=data_j.active
    h=0
    for row in data_j_s.iter_rows(min_row=8,min_col=2,max_col=2,max_row=50,values_only=True):
        h+=1
        prix=float(row[0])
        date=(j-12)*24+h
        Prix_liste.append(prix)

Prix_energie=xl.Workbook()
Prix_energie_feuille=Prix_energie.active

Prix_energie_feuille['A1']='Dates'
Prix_energie_feuille['B1']='Prix'

for h in range(len(Prix_liste)):
    Prix_energie_feuille[f'B{h+2}']=Prix_liste[h]
    Prix_energie_feuille[f'A{h+2}']=h


Prix_energie.save('Prix_Energie.xlsx')

